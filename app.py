import os
import sqlite3
import threading
from datetime import datetime, timedelta
from threading import Thread
from flask import Flask, render_template, jsonify, request, send_file, session, redirect, url_for, render_template_string
import json
from flask_socketio import SocketIO

app = Flask(__name__)
app.secret_key = "replace_with_your_secret"

# --- Session behaviour ---
app.config["SESSION_TYPE"] = "filesystem"
app.config["SESSION_PERMANENT"] = False
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=1)
app.config["SESSION_COOKIE_SAMESITE"] = "Strict"
app.config["SESSION_PERMANENT"] = False
app.config["PERMANENT_SESSION_LIFETIME"] = 0

socketio = SocketIO(app, cors_allowed_origins="*")

# SQLite database path
BASE_DIR = os.path.dirname(__file__)
DB_PATH = os.path.join(BASE_DIR, "attendance.db")

# --- Load students ---
all_students = {}

def _normalize_student(raw: dict) -> dict:
    name = (
        raw.get("name")
        or raw.get("student_name")
        or raw.get("fullName")
        or raw.get("studentName")
        or ""
    )
    section = raw.get("section") or raw.get("class") or raw.get("dept") or ""
    clazz = raw.get("section") or raw.get("class") or raw.get("dept") or "—"
    return {"name": name, "section": section, "class": clazz}


def _load_students():
    global all_students
    all_students = {}
    candidates = []
    student_dir = os.path.join(BASE_DIR, "student_data")
    if os.path.isdir(student_dir):
        for fname in os.listdir(student_dir):
            if fname.lower().endswith(".json"):
                candidates.append(os.path.join(student_dir, fname))

    for legacy in ("final_year.json", "second_year.json"):
        legacy_path = os.path.join(BASE_DIR, legacy)
        if os.path.exists(legacy_path):
            candidates.append(legacy_path)

    for path in candidates:
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict) and "students" in data and isinstance(data["students"], list):
                for s in data["students"]:
                    roll = str(s.get("roll") or s.get("roll_no") or s.get("id") or "").strip()
                    if not roll:
                        continue
                    all_students[roll.upper()] = _normalize_student(s)
            elif isinstance(data, dict):
                for k, v in data.items():
                    roll = str(k).strip()
                    if not roll:
                        continue
                    v = v if isinstance(v, dict) else {}
                    all_students[roll.upper()] = _normalize_student(v)
        except Exception as e:
            print(f"Failed to load students from {path}: {e}")


# --- Database Setup ---
def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                barcode TEXT,
                name TEXT,
                section TEXT,
                class TEXT,
                date TEXT,
                in_time TEXT,
                out_time TEXT,
                status TEXT
            )
        """)
        conn.commit()
        try:
            cur.execute("ALTER TABLE attendance ADD COLUMN section TEXT")
            conn.commit()
        except Exception:
            pass


# --- Determine In/Out ---
def determine_in_out(conn: sqlite3.Connection, barcode: str, student_name: str | None, section: str | None):
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM attendance WHERE barcode=? AND status=? ORDER BY id DESC LIMIT 1",
        (barcode, "In Library"),
    )
    row = cur.fetchone()
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    if row:  # Walk-Out
        record_id = row[0]
        cur.execute(
            "UPDATE attendance SET out_time=?, status=? WHERE id=?",
            (time_str, "Completed", record_id),
        )
        conn.commit()
        return {
            "roll": barcode,
            "barcode": barcode,
            "name": student_name or f"Student {barcode}",
            "class": "BCA",
            "section": section or "—",
            "date": date_str,
            "inTime": "",
            "outTime": time_str,
            "status": "Completed",
            "action": "Walk-Out",
        }

    # Walk-In
    cur.execute(
        "INSERT INTO attendance (barcode, name, section, class, date, in_time, status) VALUES (?,?,?,?,?,?,?)",
        (barcode, student_name or f"Student {barcode}", section or "—", (section or "—"), date_str, time_str, "In Library"),
    )
    conn.commit()
    return {
        "roll": barcode,
        "barcode": barcode,
        "name": student_name or f"Student {barcode}",
        "class": section or "—",
        "section": section or "—",
        "date": date_str,
        "inTime": time_str,
        "outTime": "—",
        "status": "In Library",
        "action": "Walk-In",
    }


# --- Real-time Summary Update Function ---
def emit_summary_update():
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        today = datetime.now().strftime("%Y-%m-%d")

        cur.execute("SELECT COUNT(*) FROM attendance WHERE date=?", (today,))
        total_walkins = cur.fetchone()[0] or 0

        cur.execute("SELECT COUNT(*) FROM attendance WHERE date=? AND status='Completed'", (today,))
        total_walkouts = cur.fetchone()[0] or 0

        cur.execute("SELECT COUNT(*) FROM attendance WHERE date=? AND status='In Library'", (today,))
        active_students = cur.fetchone()[0] or 0

    socketio.emit("update_summary", {
        "walkins": total_walkins,
        "walkouts": total_walkouts,
        "active": active_students
    })


# --- Routes ---
@app.route("/")
def index():
    return render_template("index.html")


@app.get("/api/attendance")
def get_attendance():
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT barcode, name, section, class, date, in_time, out_time, status FROM attendance ORDER BY id DESC LIMIT 1000"
        )
        rows = cur.fetchall()
        data = [
            {
                "roll": r[0],
                "barcode": r[0],
                "name": r[1],
                "section": r[2],
                "class": r[3],
                "date": r[4],
                "inTime": r[5],
                "outTime": r[6] if r[6] else "—",
                "status": r[7],
            }
            for r in rows
        ]
      
        return jsonify({"attendance": list(reversed(data))})


@app.route('/api/clear_attendance', methods=['DELETE'])
def clear_attendance():
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("DELETE FROM attendance")
        conn.commit()
        conn.close()
        emit_summary_update()  # 🔄 refresh after clearing
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# --- Barcode Callback ---
def on_barcode(barcode_value: str):
    barcode_value = (barcode_value or "").strip()
    if not barcode_value:
        return
    with sqlite3.connect(DB_PATH) as conn:
        student = all_students.get(barcode_value.upper()) or {}
        student_name = (student.get("name") or "").strip() or None
        section = (student.get("section") or "").strip() or None
        record = determine_in_out(conn, barcode_value, student_name, section)
    payload = {
        "roll_no": barcode_value,
        "student_name": record.get("name") if student_name else "Student not found",
        "section": section or "—",
        "action": record.get("action"),
        # Prefer outTime only if it's a real timestamp; otherwise use inTime
        "time": (record.get("outTime") if record.get("outTime") not in (None, "", "—") else record.get("inTime")),
    }
    socketio.emit("barcode_scanned", payload)
    emit_summary_update()  # 🔄 emit live summary update


# --- Background Scanner ---
_scanner_thread = None

def start_barcode_listener_background():
    global _scanner_thread
    if _scanner_thread and _scanner_thread.is_alive():
        return False

    from both_test import main_listener

    def run():
        try:
            main_listener(on_barcode)
        except Exception as e:
            print(f"Scanner listener exited: {e}")

    _scanner_thread = Thread(target=run, daemon=True)
    _scanner_thread.start()
    return True


@app.post("/api/start_scanner")
def api_start_scanner():
    started = start_barcode_listener_background()
    return jsonify({"success": True, "started": started})


# --- Admin Login ---
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "1"

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session["logged_in"] = True
            return redirect(url_for("index"))
        else:
            return render_template_string(LOGIN_TEMPLATE, error="Invalid credentials")
    return render_template_string(LOGIN_TEMPLATE, error=None)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.before_request
def require_login():
    open_routes = (
        "login",
        "static",
        "api_start_scanner",
        "get_attendance",
        "export_excel",
        "export_pdf",
        "clear_attendance",
    )
    if request.endpoint in open_routes or request.endpoint is None:
        return
    if not session.get("logged_in"):
        return redirect(url_for("login"))


LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <title>Admin Login</title>
  <script src="https://cdn.tailwindcss.com"></script>
</head>
<body class="bg-slate-100 flex items-center justify-center h-screen">
  <div class="bg-white p-8 rounded-2xl shadow-md w-80">
    <h1 class="text-xl font-bold text-slate-800 mb-4 text-center">Admin Login</h1>
    {% if error %}
    <p class="text-red-600 text-sm mb-3 text-center">{{ error }}</p>
    {% endif %}
    <form method="POST">
      <label class="block text-sm text-slate-600 mb-1">Username</label>
      <input type="text" name="username" class="w-full border border-slate-300 rounded-lg p-2 mb-3 focus:outline-none focus:ring-2 focus:ring-emerald-500">
      <label class="block text-sm text-slate-600 mb-1">Password</label>
      <input type="password" name="password" class="w-full border border-slate-300 rounded-lg p-2 mb-4 focus:outline-none focus:ring-2 focus:ring-emerald-500">
      <button class="w-full bg-emerald-600 hover:bg-emerald-700 text-white py-2 rounded-lg font-semibold">Login</button>
    </form>
  </div>
</body>
</html>
"""


# --- Exports ---
@app.get("/export/excel")
def export_excel():
    # Try to generate a real XLSX; if openpyxl is missing, fall back to CSV
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Detect filters
        department = (request.args.get("department") or "").strip()
        start_date = (request.args.get("startDate") or "").strip()
        end_date = (request.args.get("endDate") or "").strip()
        is_single_class = bool(department and department.lower() != "all")
        is_single_date = bool(start_date and start_date == end_date)

        # Build headers conditionally
        headers = ["Roll", "Name"]
        if not is_single_class:
            headers.append("Class")
        if not is_single_date:
            headers.append("Date")
        headers.extend(["Walk-In Time", "Walk-Out Time", "Status"])

        # Standalone info box above the header (merged single cell with border and wrapping)
        info_parts = []
        if is_single_class:
            info_parts.append(f"Class: {department}")
        if is_single_date:
            info_parts.append(f"Date: {start_date}")
        info_text = "\n".join(info_parts)
        if info_text:
            total_cols = len(headers)
            ws.append([info_text])
            r = ws.max_row
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
            try:
                from openpyxl.styles import Border, Side
                thin = Side(style="thin")
                box_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            except Exception:
                box_border = None
            cell = ws.cell(row=r, column=1)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if box_border:
                cell.border = box_border
            longest = max((len(p) for p in info_parts), default=0)
            approx_lines = max(1, len(info_parts) + longest // 40)
            ws.row_dimensions[r].height = 18 * approx_lines

        ws.append(headers)
        bold = Font(bold=True)
        header_row = ws.max_row
        total_cols = len(headers)
        for idx in range(1, total_cols + 1):
            ws.cell(row=header_row, column=idx).font = bold

        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            # Apply optional filters from query string
            clauses = []
            params = []
            department = (request.args.get("department") or "").strip()
            start_date = (request.args.get("startDate") or "").strip()
            end_date = (request.args.get("endDate") or "").strip()

            if department and department.lower() != "all":
                clauses.append("(LOWER(class) LIKE ? OR LOWER(section) LIKE ?)")
                like = f"%{department.lower()}%"
                params.extend([like, like])
            if start_date:
                clauses.append("date >= ?")
                params.append(start_date)
            if end_date:
                clauses.append("date <= ?")
                params.append(end_date)

            where_sql = (" WHERE " + " AND ".join(clauses)) if clauses else ""
            sql = (
                "SELECT barcode, name, section, class, date, in_time, out_time, status FROM attendance"
                + where_sql + " ORDER BY id ASC"
            )
            cur.execute(sql, params)
            for r in cur.fetchall():
                row = [
                    r[0] or "",
                    r[1] or "",
                ]
                if not is_single_class:
                    row.append(r[3] or "")
                if not is_single_date:
                    row.append(r[4] or "")
                row.extend([
                    r[5] or "",
                    (r[6] or "—"),
                    r[7] or "",
                ])
                ws.append(row)

        # Formatting: freeze header, wrap, auto-width based on header+data only
        from openpyxl.utils import get_column_letter
        ws.freeze_panes = f"A{header_row + 1}"
        last_row = ws.max_row
        for col_idx in range(1, total_cols + 1):
            max_len = 0
            for row in ws.iter_rows(min_row=header_row, max_row=last_row, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            width = max(12, min(60, int(max_len * 1.2)))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        bio = io.BytesIO()
        wb.save(bio)
        data_bytes = bio.getvalue()
        # Safety check: XLSX should be a ZIP file starting with 'PK' signature
        if not data_bytes or not data_bytes.startswith(b"PK"):
            raise RuntimeError("Generated XLSX failed integrity check; falling back to CSV")
        bio.seek(0)
        return send_file(
            bio,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="attendance.xlsx",
        )
    except Exception:
        import io, csv
        output = io.StringIO()
        writer = csv.writer(output)
        department = (request.args.get("department") or "").strip()
        start_date = (request.args.get("startDate") or "").strip()
        end_date = (request.args.get("endDate") or "").strip()
        is_single_class = bool(department and department.lower() != "all")
        is_single_date = bool(start_date and start_date == end_date)

        if is_single_class:
            writer.writerow([f"Class: {department}"])
        if is_single_date:
            writer.writerow([f"Date: {start_date}"])

        headers = ["Roll", "Name"]
        if not is_single_class:
            headers.append("Class")
        if not is_single_date:
            headers.append("Date")
        headers.extend(["Walk-In Time", "Walk-Out Time", "Status"])
        writer.writerow(headers)
        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            # Apply optional filters from query string
            clauses = []
            params = []
            department = (request.args.get("department") or "").strip()
            start_date = (request.args.get("startDate") or "").strip()
            end_date = (request.args.get("endDate") or "").strip()

            if department and department.lower() != "all":
                clauses.append("(LOWER(class) LIKE ? OR LOWER(section) LIKE ?)")
                like = f"%{department.lower()}%"
                params.extend([like, like])
            if start_date:
                clauses.append("date >= ?")
                params.append(start_date)
            if end_date:
                clauses.append("date <= ?")
                params.append(end_date)

            where_sql = (" WHERE " + " AND ".join(clauses)) if clauses else ""
            sql = (
                "SELECT barcode, name, section, class, date, in_time, out_time, status FROM attendance"
                + where_sql + " ORDER BY id ASC"
            )
            cur.execute(sql, params)
            for r in cur.fetchall():
                row = [
                    r[0] or "",
                    r[1] or "",
                ]
                if not is_single_class:
                    row.append(r[3] or "")
                if not is_single_date:
                    row.append(r[4] or "")
                row.extend([
                    r[5] or "",
                    (r[6] or "—"),
                    r[7] or "",
                ])
                writer.writerow(row)
        data = output.getvalue().encode("utf-8-sig")
        mem = io.BytesIO(data)
        mem.seek(0)
        return send_file(
            mem,
            mimetype="text/csv; charset=utf-8",
            as_attachment=True,
            download_name="attendance.csv",
        )


@app.get("/export/pdf")
def export_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
    except Exception:
        return jsonify({
            "success": False,
            "error": "PDF export not configured. Install reportlab or use CSV export.",
            "hint": "pip install reportlab"
        }), 501

    import io
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Margins
    left_margin = 1.5*cm
    right_margin = 1.5*cm

    # Header with logo + college title
    from reportlab.lib.utils import ImageReader
    import os as _os
    logo_path = _os.path.join(BASE_DIR, "static", "logo.jpg")

    def draw_header() -> float:
        top_y = height - 1.2*cm
        logo_w = 1.8*cm
        logo_h = 1.8*cm
        has_logo = _os.path.exists(logo_path)
        if has_logo:
            try:
                c.drawImage(logo_path, left_margin, top_y - logo_h, width=logo_w, height=logo_h, preserveAspectRatio=True, mask='auto')
            except Exception:
                has_logo = False
        text_x = left_margin + (logo_w if has_logo else 0) + 0.5*cm
        # Title
        c.setFont("Helvetica-Bold", 16)
        c.drawString(text_x, top_y - 0.2*cm, "Dr. B. B. Hegde First Grade College, Kundapura")
        # Subtitle
        c.setFont("Helvetica", 10)
        c.drawString(text_x, top_y - 0.2*cm - 0.7*cm, "A Unit of Coondapur Education Society (R)")
        # underline
        line_y = top_y - logo_h - 0.25*cm
        c.setLineWidth(0.5)
        c.line(left_margin, line_y, width - right_margin, line_y)
        # Optional report title below
        c.setFont("Helvetica-Bold", 12)
        c.drawString(left_margin, line_y - 0.9*cm, "Library Attendance Report")
        return line_y - 1.1*cm

    top_start = draw_header()
    bottom_margin = 2*cm
    usable_width = width - left_margin - right_margin

    x0 = left_margin
    y = top_start
    row_height = 0.65*cm

    # Detect filters to conditionally show Class/Date and draw info box
    department = (request.args.get("department") or "").strip()
    start_date = (request.args.get("startDate") or "").strip()
    end_date = (request.args.get("endDate") or "").strip()
    is_single_class = bool(department and department.lower() != "all")
    is_single_date = bool(start_date and start_date == end_date)

    headers = ["Roll", "Name"]
    if not is_single_class:
        headers.append("Class")
    if not is_single_date:
        headers.append("Date")
    headers.extend(["In", "Out", "Status"])

    def draw_row(values, is_header=False):
        nonlocal y
        # Draw background for header
        if is_header:
            c.setLineWidth(1)
            c.rect(x0, y - row_height, total_cols_width, row_height, stroke=1, fill=0)
            c.setFont("Helvetica-Bold", 9)
        else:
            c.setLineWidth(0.5)
            c.rect(x0, y - row_height, total_cols_width, row_height, stroke=1, fill=0)
            c.setFont("Helvetica", 9)

        # Vertical lines and cell text
        x = x0
        for i, (text, cw) in enumerate(zip(values, col_widths)):
            # Cell box
            if i > 0:
                c.line(x, y - row_height, x, y)
            # Text clipped to cell
            clip = str(text or "")
            # Simple clip: reduce until it fits
            max_chars = int(cw / 5.5)  # heuristic width per char
            if len(clip) > max_chars:
                clip = clip[:max_chars-1] + "…"
            c.drawString(x + 2, y - row_height + 2, clip)
            x += cw

        y -= row_height

    # Compute column widths now that we know which headers are present
    if not is_single_class and not is_single_date:
        col_widths_cm = [3.0, 7.0, 3.5, 3.0, 2.5, 2.5, 3.0]
    elif is_single_class and not is_single_date:
        col_widths_cm = [3.0, 8.5, 3.5, 2.5, 2.5, 3.0]  # no Class column
    elif not is_single_class and is_single_date:
        col_widths_cm = [3.0, 8.5, 3.5, 2.5, 2.5, 3.0]  # no Date column
    else:
        col_widths_cm = [3.0, 10.0, 2.8, 2.8, 3.0]  # no Class, no Date
    col_widths = [w*cm for w in col_widths_cm]
    total_cols_width = sum(col_widths)
    if total_cols_width > usable_width:
        scale = usable_width / total_cols_width
        col_widths = [w*scale for w in col_widths]
        total_cols_width = sum(col_widths)

    # Optional info box (separate bordered rectangle with wrapped text)
    if is_single_class or is_single_date:
        info_lines = []
        if is_single_class:
            info_lines.append(f"Class: {department}")
        if is_single_date:
            info_lines.append(f"Date: {start_date}")
        info_text = "\n".join(info_lines)
        # Box geometry
        box_height = max(1.0*cm, 0.6*cm * len(info_lines) + 0.4*cm)
        c.setLineWidth(0.8)
        c.rect(x0, y - box_height, total_cols_width, box_height, stroke=1, fill=0)
        c.setFont("Helvetica-Bold", 10)
        text_y = y - 0.35*cm
        for line in info_lines:
            c.drawString(x0 + 0.2*cm, text_y, line)
            text_y -= 0.6*cm
        y -= (box_height + 0.25*cm)

    # Header row
    draw_row(headers, is_header=True)

    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        # Apply optional filters from query string
        clauses = []
        params = []
        department = (request.args.get("department") or "").strip()
        start_date = (request.args.get("startDate") or "").strip()
        end_date = (request.args.get("endDate") or "").strip()

        if department and department.lower() != "all":
            clauses.append("(LOWER(class) LIKE ? OR LOWER(section) LIKE ?)")
            like = f"%{department.lower()}%"
            params.extend([like, like])
        if start_date:
            clauses.append("date >= ?")
            params.append(start_date)
        if end_date:
            clauses.append("date <= ?")
            params.append(end_date)

        where_sql = (" WHERE " + " AND ".join(clauses)) if clauses else ""
        sql = (
            "SELECT barcode, name, section, class, date, in_time, out_time, status FROM attendance"
            + where_sql + " ORDER BY id ASC"
        )
        cur.execute(sql, params)
        for r in cur.fetchall():
            # Build row conditionally (omit repeated Class/Date)
            row = [
                str(r[0] or ""), str(r[1] or ""),
            ]
            if not is_single_class:
                row.append(str(r[3] or ""))
            if not is_single_date:
                row.append(str(r[4] or ""))
            row.extend([
                str(r[5] or ""), str(r[6] or "—"), str(r[7] or "")
            ])
            # New page if needed
            if y - row_height < bottom_margin:
                c.showPage()
                c.setFont("Helvetica", 9)
                # redraw header on each new page
                new_top = draw_header()
                y = new_top
                if is_single_class:
                    draw_row([f"Class: {department}"] + [""] * (len(headers) - 1))
                if is_single_date:
                    draw_row([f"Date: {start_date}"] + [""] * (len(headers) - 1))
                draw_row(headers, is_header=True)
            draw_row(row)

    c.showPage()
    c.save()

    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="attendance.pdf",
    )


if __name__ == "__main__":
    _load_students()
    init_db()
    start_barcode_listener_background()
    socketio.run(app, host="0.0.0.0", port=5001)